"""
Excel Report Writer for Slab Take-Off
======================================
Generates a formatted Excel workbook with:
  - Sheet 1: "Slab Details" — every slab with its quantities
  - Sheet 2: "Summary" — totals grouped by thickness
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os

from slab_extractor import SlabReport
from wall_extractor import WallReport
from beam_extractor import BeamReport


# ── Style constants ──────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=10, color="666666")
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
WARNING_FONT = Font(name="Calibri", size=10, color="CC0000")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def write_excel_report(report: SlabReport, output_path: str, image_path: str = None, 
                       wall_report: WallReport = None, str_wall_img: str = None, ns_wall_img: str = None,
                       beam_report: BeamReport = None, beam_img: str = None):
    """
    Write the slab and wall take-off report to a formatted Excel file.
    """
    wb = Workbook()
    
    _write_details_sheet(wb, report)
    _write_summary_sheet(wb, report)
    
    if image_path and os.path.exists(image_path):
        _write_map_sheet(wb, "Slab Map", "SLAB VISUAL MAP", image_path)
        
    if wall_report:
        _write_wall_sheet(wb, "Structural Walls", wall_report.structural_walls, wall_report.structural_summary, wall_report, str_wall_img)
        _write_wall_sheet(wb, "Non-Structural Walls", wall_report.non_structural_walls, wall_report.non_structural_summary, wall_report, ns_wall_img)

    if beam_report:
        _write_beam_sheet(wb, beam_report, beam_img)

    wb.save(output_path)

def _write_wall_sheet(wb: Workbook, sheet_name: str, walls, summary, report: WallReport, img_path: str):
    ws = wb.create_sheet(sheet_name)
    
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{sheet_name.upper()} QUANTITY TAKE-OFF"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT
    
    # ── Interactive Height Input ────────────────────────────────
    ws["A3"] = "Floor-to-Floor Height (m):"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="1F3864")
    
    height_cell = ws["B3"]
    height_cell.value = 3.0  # Default value, user can change this!
    height_cell.font = Font(name="Calibri", bold=True, size=12, color="000000")
    height_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # yellow highlight
    height_cell.alignment = CENTER
    height_cell.border = THIN_BORDER
    
    # ── Header row ────────────────────────────────────────────
    headers = [("Wall Name", 14), ("Layer", 22), ("Thickness (mm)", 16), 
               ("Length (m)", 14), ("Area (m\u00b2)", 14), ("Volume (m\u00b3)", 14), 
               ("Centroid", 24)]
    
    header_row = 5
    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # ── Data rows ─────────────────────────────────────────────
    for i, w in enumerate(walls):
        row = header_row + 1 + i
        fill = ALT_ROW_FILL if i % 2 == 0 else None
        
        values = [w.name, w.layer, w.thickness_mm, w.length_m, w.area_sqm, None, f"({w.centroid_x:.0f}, {w.centroid_y:.0f})"]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill: cell.fill = fill
            
            if col_idx in (3,): cell.alignment = CENTER
            elif col_idx in (4, 5, 6): 
                cell.alignment = RIGHT
                cell.number_format = "#,##0.000"
                if col_idx == 6: # Volume Formula: Area * Height (B3)
                    area_col = get_column_letter(5)
                    cell.value = f"={area_col}{row}*B$3"
            else: cell.alignment = LEFT
            
    # ── Grand total row ───────────────────────────────────────
    total_row = header_row + 1 + len(walls)
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    
    for col in range(2, 4):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER
        
    ws.cell(row=total_row, column=4, value=f"=SUM(D{header_row+1}:D{total_row-1})").font = TOTAL_FONT
    ws.cell(row=total_row, column=4).fill = TOTAL_FILL
    ws.cell(row=total_row, column=4).number_format = "#,##0.000"
    ws.cell(row=total_row, column=4).border = THIN_BORDER
    
    ws.cell(row=total_row, column=5, value=f"=SUM(E{header_row+1}:E{total_row-1})").font = TOTAL_FONT
    ws.cell(row=total_row, column=5).fill = TOTAL_FILL
    ws.cell(row=total_row, column=5).number_format = "#,##0.000"
    ws.cell(row=total_row, column=5).border = THIN_BORDER
    
    ws.cell(row=total_row, column=6, value=f"=SUM(F{header_row+1}:F{total_row-1})").font = TOTAL_FONT
    ws.cell(row=total_row, column=6).fill = TOTAL_FILL
    ws.cell(row=total_row, column=6).number_format = "#,##0.000"
    ws.cell(row=total_row, column=6).border = THIN_BORDER

    ws.cell(row=total_row, column=7).fill = TOTAL_FILL
    ws.cell(row=total_row, column=7).border = THIN_BORDER
    
    if img_path and os.path.exists(img_path):
        img = Image(img_path)
        img.width = img.width * 0.20
        img.height = img.height * 0.20
        ws.add_image(img, f"I{header_row}")
    
    ws.freeze_panes = f"A{header_row + 1}"

def _write_details_sheet(wb: Workbook, report: SlabReport):
    """Sheet 1: Every slab with its quantities."""
    ws = wb.active
    ws.title = "Slab Details"
    
    # ── Title block ───────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = f"SLAB QUANTITY TAKE-OFF"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT
    
    ws.merge_cells("A2:H2")
    ws["A2"] = f"File: {report.file_name}  |  Generated: {report.generated_at}"
    ws["A2"].font = SUBTITLE_FONT
    
    ws.merge_cells("A3:H3")
    ws["A3"] = (
        f"Total slabs: {report.total_slabs}  |  "
        f"Matched by text label: {report.matched_by_text}  |  "
        f"Matched by layer name: {report.matched_by_layer}"
    )
    ws["A3"].font = SUBTITLE_FONT
    
    # ── Header row ────────────────────────────────────────────
    headers = [
        ("Slab Name", 14),
        ("Layer", 22),
        ("Thickness (mm)", 16),
        ("Source", 14),
        ("Label Text", 16),
        ("Area (m\u00b2)", 14),
        ("Volume (m\u00b3)", 14),
        ("Centroid", 24),
    ]
    
    header_row = 5
    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ── Data rows ─────────────────────────────────────────────
    for i, slab in enumerate(report.slabs):
        row = header_row + 1 + i
        fill = ALT_ROW_FILL if i % 2 == 0 else None
        
        values = [
            slab.name,
            slab.layer,
            slab.thickness_mm,
            slab.thickness_source,
            slab.label_text,
            slab.area_sqm,
            slab.volume_cum,
            f"({slab.centroid_x:.0f}, {slab.centroid_y:.0f})",
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            
            # Alignment
            if col_idx in (3,):  # thickness
                cell.alignment = CENTER
            elif col_idx in (6, 7):  # area, volume
                cell.alignment = RIGHT
                cell.number_format = "#,##0.000"
            elif col_idx == 4:  # source
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
    
    # ── Grand total row ───────────────────────────────────────
    total_row = header_row + 1 + len(report.slabs)
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    
    for col in range(2, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    
    ws.cell(row=total_row, column=5, value=f"{report.total_slabs} slabs").font = TOTAL_FONT
    ws.cell(row=total_row, column=5).fill = TOTAL_FILL
    ws.cell(row=total_row, column=5).alignment = RIGHT
    ws.cell(row=total_row, column=5).border = THIN_BORDER
    
    area_cell = ws.cell(row=total_row, column=6, value=report.grand_total_area)
    area_cell.font = TOTAL_FONT
    area_cell.fill = TOTAL_FILL
    area_cell.alignment = RIGHT
    area_cell.number_format = "#,##0.000"
    area_cell.border = THIN_BORDER
    
    vol_cell = ws.cell(row=total_row, column=7, value=report.grand_total_volume)
    vol_cell.font = TOTAL_FONT
    vol_cell.fill = TOTAL_FILL
    vol_cell.alignment = RIGHT
    vol_cell.number_format = "#,##0.000"
    vol_cell.border = THIN_BORDER
    
    ws.cell(row=total_row, column=8).fill = TOTAL_FILL
    ws.cell(row=total_row, column=8).border = THIN_BORDER
    
    # ── Warnings section ──────────────────────────────────────
    if report.warnings:
        warn_start = total_row + 2
        ws.cell(row=warn_start, column=1, value="WARNINGS:").font = Font(
            name="Calibri", bold=True, size=11, color="CC0000"
        )
        for j, warning in enumerate(report.warnings):
            ws.cell(row=warn_start + 1 + j, column=1, value=f"  - {warning}").font = WARNING_FONT
    
    # Freeze panes: header row stays visible while scrolling
    ws.freeze_panes = f"A{header_row + 1}"


def _write_summary_sheet(wb: Workbook, report: SlabReport):
    """Sheet 2: Summary by thickness."""
    ws = wb.create_sheet("Summary")
    
    # ── Title ─────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = "SLAB QUANTITY SUMMARY"
    ws["A1"].font = TITLE_FONT
    
    ws.merge_cells("A2:D2")
    ws["A2"] = f"File: {report.file_name}  |  Generated: {report.generated_at}"
    ws["A2"].font = SUBTITLE_FONT
    
    # ── Headers ───────────────────────────────────────────────
    headers = [
        ("Slab Thickness", 20),
        ("Count", 12),
        ("Total Area (m\u00b2)", 18),
        ("Total Volume (m\u00b3)", 18),
    ]
    
    header_row = 4
    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ── Data ──────────────────────────────────────────────────
    for i, row_data in enumerate(report.summary):
        row = header_row + 1 + i
        fill = ALT_ROW_FILL if i % 2 == 0 else None
        
        values = [
            f"{row_data.thickness_mm} mm",
            row_data.count,
            row_data.total_area_sqm,
            row_data.total_volume_cum,
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if col_idx == 1:
                cell.alignment = LEFT
            elif col_idx == 2:
                cell.alignment = CENTER
            else:
                cell.alignment = RIGHT
                cell.number_format = "#,##0.000"
    
    # ── Grand total ───────────────────────────────────────────
    total_row = header_row + 1 + len(report.summary)
    
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    
    count_cell = ws.cell(row=total_row, column=2, value=report.total_slabs)
    count_cell.font = TOTAL_FONT
    count_cell.fill = TOTAL_FILL
    count_cell.alignment = CENTER
    count_cell.border = THIN_BORDER
    
    area_cell = ws.cell(row=total_row, column=3, value=report.grand_total_area)
    area_cell.font = TOTAL_FONT
    area_cell.fill = TOTAL_FILL
    area_cell.alignment = RIGHT
    area_cell.number_format = "#,##0.000"
    area_cell.border = THIN_BORDER
    
    vol_cell = ws.cell(row=total_row, column=4, value=report.grand_total_volume)
    vol_cell.font = TOTAL_FONT
    vol_cell.fill = TOTAL_FILL
    vol_cell.alignment = RIGHT
    vol_cell.number_format = "#,##0.000"
    vol_cell.border = THIN_BORDER
    
    ws.freeze_panes = f"A{header_row + 1}"


def _write_map_sheet(wb: Workbook, sheet_title: str, heading: str, image_path: str):
    """Sheet 3: Visual Map."""
    ws = wb.create_sheet(sheet_title)
    
    ws.merge_cells("A1:H1")
    ws["A1"] = heading
    ws["A1"].font = TITLE_FONT
    
    ws.merge_cells("A2:H2")
    ws["A2"] = "Color-coded by thickness. Labels placed at slab centroids."
    ws["A2"].font = SUBTITLE_FONT
    
    img = Image(image_path)
    # The image is large, we can scale it down slightly to fit nicely in Excel
    # ezdxf matplotlib creates a large figure by default (e.g., 24x17 @ 200dpi is ~4800x3400 px)
    # Let's scale it to fit within typical screen bounds
    img.width = img.width * 0.25
    img.height = img.height * 0.25
    
    ws.add_image(img, "A4")

def _write_beam_sheet(wb: Workbook, report: BeamReport, img_path: str):
    ws = wb.create_sheet("Beams")
    
    ws.merge_cells("A1:H1")
    ws["A1"] = "BEAM QUANTITY TAKE-OFF"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT
    
    ws.merge_cells("A2:H2")
    ws["A2"] = "Note: Length shown is Clear Span (Structural overlaps subtracted)."
    ws["A2"].font = SUBTITLE_FONT
    
    # ── Header row ────────────────────────────────────────────
    headers = [("Beam Name", 14), ("Width (mm)", 14), ("Depth (mm)", 14), 
               ("CAD Span (m)", 14), ("Clear Span (m)", 14), ("Volume (m\u00b3)", 14), 
               ("Centroid", 24)]
    
    header_row = 4
    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # ── Data rows ─────────────────────────────────────────────
    for i, b in enumerate(report.beams):
        row = header_row + 1 + i
        fill = ALT_ROW_FILL if i % 2 == 0 else None
        
        values = [b.name, b.width*1000, b.depth*1000, b.cad_length, b.clear_length, b.volume, f"({b.cx:.0f}, {b.cy:.0f})"]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill: cell.fill = fill
            
            if col_idx in (2,3): cell.alignment = CENTER
            elif col_idx in (4,5,6): 
                cell.alignment = RIGHT
                cell.number_format = "#,##0.000"
            else: cell.alignment = LEFT
            
    # ── Grand total row ───────────────────────────────────────
    total_row = header_row + 1 + len(report.beams)
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    
    for col in range(2, 6):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER
        
    ws.cell(row=total_row, column=6, value=f"=SUM(F{header_row+1}:F{total_row-1})").font = TOTAL_FONT
    ws.cell(row=total_row, column=6).fill = TOTAL_FILL
    ws.cell(row=total_row, column=6).number_format = "#,##0.000"
    ws.cell(row=total_row, column=6).border = THIN_BORDER

    ws.cell(row=total_row, column=7).fill = TOTAL_FILL
    ws.cell(row=total_row, column=7).border = THIN_BORDER
    
    if img_path and os.path.exists(img_path):
        img = Image(img_path)
        img.width = img.width * 0.20
        img.height = img.height * 0.20
        ws.add_image(img, f"I{header_row}")
    
    ws.freeze_panes = f"A{header_row + 1}"
