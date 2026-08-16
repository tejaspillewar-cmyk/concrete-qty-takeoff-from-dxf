import sys
import os
import threading
import tempfile
import tkinter as tk
from tkinter import filedialog
import customtkinter as ctk

# Add the 01_slab_takeoff directory to path so we can import its modules
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SLAB_DIR = os.path.join(BASE_DIR, "01_slab_takeoff")
sys.path.insert(0, SLAB_DIR)

from slab_extractor import extract_slabs
from excel_report import write_excel_report
from slab_visualizer import render_slab_map
from wall_extractor import extract_walls
from wall_visualizer import render_wall_map

# Set appearance mode and color theme
ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

class TakeoffApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Slab Quantity Take-Off")
        self.geometry("1000x650")
        
        # grid layout 1x2 (Left: Controls, Right: Preview)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # ── Left Frame (Controls) ──────────────────────────────────
        self.left_frame = ctk.CTkFrame(self)
        self.left_frame.grid(row=0, column=0, padx=(20, 10), pady=20, sticky="nsew")
        
        self.left_frame.grid_columnconfigure(0, weight=1)
        self.left_frame.grid_rowconfigure(4, weight=1)

        # ── Right Frame (Preview) ──────────────────────────────────
        self.right_frame = ctk.CTkFrame(self)
        self.right_frame.grid(row=0, column=1, padx=(10, 20), pady=20, sticky="nsew")
        self.right_frame.grid_rowconfigure(0, weight=1)
        self.right_frame.grid_columnconfigure(0, weight=1)
        
        self.image_label = ctk.CTkLabel(self.right_frame, text="Map Preview Will Appear Here", text_color="gray")
        self.image_label.grid(row=0, column=0, sticky="nsew")
        
        # grid layout 1x2
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)



        # ── Header ──────────────────────────────────────────────────
        self.title_label = ctk.CTkLabel(self.left_frame, text="Slab Quantity Take-Off", font=ctk.CTkFont(size=24, weight="bold"))
        self.title_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.subtitle_label = ctk.CTkLabel(self.left_frame, text="Extract structural slab areas and volumes from DXF files.", font=ctk.CTkFont(size=14))
        self.subtitle_label.grid(row=1, column=0, padx=20, pady=(0, 20))

        # ── File Selection ──────────────────────────────────────────
        self.file_frame = ctk.CTkFrame(self.left_frame, fg_color="transparent")
        self.file_frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        self.file_frame.grid_columnconfigure(1, weight=1)

        self.select_btn = ctk.CTkButton(self.file_frame, text="Select DXF File", command=self.select_file)
        self.select_btn.grid(row=0, column=0, padx=(0, 10))

        self.file_path_var = tk.StringVar(value="No file selected")
        self.file_path_entry = ctk.CTkEntry(self.file_frame, textvariable=self.file_path_var, state="readonly")
        self.file_path_entry.grid(row=0, column=1, sticky="ew")

        # ── Settings ────────────────────────────────────────────────
        self.settings_frame = ctk.CTkFrame(self.left_frame, fg_color="transparent")
        self.settings_frame.grid(row=3, column=0, padx=20, pady=10, sticky="ew")
        
        self.height_label = ctk.CTkLabel(self.settings_frame, text="Floor-to-Floor Height (m):")
        self.height_label.grid(row=0, column=0, padx=(0, 10))
        
        self.height_var = tk.StringVar(value="3.0")
        self.height_entry = ctk.CTkEntry(self.settings_frame, textvariable=self.height_var, width=60)
        self.height_entry.grid(row=0, column=1)

        # ── Action ──────────────────────────────────────────────────
        self.run_btn = ctk.CTkButton(self.left_frame, text="Run Take-Off", font=ctk.CTkFont(weight="bold"), 
                                     command=self.run_takeoff, fg_color="#2FA572", hover_color="#1F7A52", height=40)
        self.run_btn.grid(row=4, column=0, padx=20, pady=20)

        # ── Output Log ──────────────────────────────────────────────
        self.log_textbox = ctk.CTkTextbox(self.left_frame, width=250)
        self.log_textbox.grid(row=5, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.log_textbox.insert("0.0", "Welcome! Select a DXF file to begin.\n")
        self.log_textbox.configure(state="disabled")

        self.selected_file = None

    def log(self, message):
        """Append a message to the UI log box."""
        self.log_textbox.configure(state="normal")
        self.log_textbox.insert("end", message + "\n")
        self.log_textbox.see("end")
        self.log_textbox.configure(state="disabled")

    def select_file(self):
        filename = filedialog.askopenfilename(
            title="Select DXF",
            filetypes=(("DXF Files", "*.dxf"), ("All Files", "*.*"))
        )
        if filename:
            self.selected_file = filename
            self.file_path_var.set(filename)
            self.log(f"Selected file: {filename}")
            
            # Generate preview in background
            self.log("Generating DXF preview...")
            self.image_label.configure(text="Loading preview...")
            threading.Thread(target=self._generate_preview, args=(filename,), daemon=True).start()

    def _generate_preview(self, dxf_path):
        try:
            import ezdxf
            from ezdxf.addons.drawing import Frontend, RenderContext
            from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
            import matplotlib
            matplotlib.use("Agg")  # Non-interactive backend
            import matplotlib.pyplot as plt

            # Read DXF and setup plot
            doc = ezdxf.readfile(dxf_path)
            msp = doc.modelspace()
            
            # Use a dark background to match the UI
            fig = plt.figure(figsize=(8, 8), facecolor="#2b2b2b")
            ax = fig.add_axes([0, 0, 1, 1])
            ax.set_facecolor("#2b2b2b")
            
            # Render
            ctx = RenderContext(doc)
            out = MatplotlibBackend(ax)
            Frontend(ctx, out).draw_layout(msp, finalize=True)
            
            # Save to temporary file
            preview_path = os.path.join(tempfile.gettempdir(), "dxf_preview.png")
            fig.savefig(preview_path, dpi=100, facecolor=fig.get_facecolor())
            plt.close(fig)
            
            # Display
            self.after(0, self._display_image, preview_path)
            self.log("Preview ready.")
        except Exception as e:
            self.log(f"Failed to generate preview: {e}")
            self.after(0, lambda: self.image_label.configure(text="Preview generation failed"))

    def run_takeoff(self):
        if not self.selected_file:
            self.log("ERROR: Please select a file first.")
            return

        # Disable buttons during run
        self.select_btn.configure(state="disabled")
        self.run_btn.configure(state="disabled")
        
        self.log("\n" + "="*50)
        self.log(f"Scanning layers for: {os.path.basename(self.selected_file)}")
        
        try:
            import ezdxf
            doc = ezdxf.readfile(self.selected_file)
            all_layers = [layer.dxf.name for layer in doc.layers]
        except Exception as e:
            self.log(f"Failed to read DXF layers: {e}")
            self._enable_buttons()
            return
            
        slab_layers = [l for l in all_layers if l.upper().startswith("STR-SLAB-REG")]
        str_wall_layers = [l for l in all_layers if "STR-WALL-REG" in l.upper()]
        ns_wall_layers = [l for l in all_layers if "NS-WALL-REG" in l.upper()]
        beam_layers = [l for l in all_layers if "BEAM" in l.upper()]
        
        self._show_layer_confirmation(slab_layers, str_wall_layers, ns_wall_layers, beam_layers)

    def _show_layer_confirmation(self, slab_layers, str_wall_layers, ns_wall_layers, beam_layers):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Confirm Detected Layers")
        dialog.geometry("600x700")
        dialog.transient(self) # dialog appears on top of main window
        dialog.grab_set()      # makes dialog modal
        
        # Add scrollable frame
        scroll = ctk.CTkScrollableFrame(dialog)
        scroll.pack(fill="both", expand=True, padx=20, pady=20)
        
        
        ctk.CTkLabel(scroll, text="Edit the layers below (one per line).", text_color="gray").pack(pady=(0, 10))
        
        self.layer_inputs = {}
        
        def add_section(title, key, layers):
            lbl = ctk.CTkLabel(scroll, text=title, font=ctk.CTkFont(weight="bold", size=14))
            lbl.pack(anchor="w", pady=(10, 5))
            
            txt = ctk.CTkTextbox(scroll, height=100)
            txt.pack(fill="x")
            txt.insert("0.0", "\n".join(layers))
            self.layer_inputs[key] = txt
                    
        add_section("Slab Layers", "slabs", slab_layers)
        add_section("Structural Wall Layers", "str_walls", str_wall_layers)
        add_section("Non-Structural Wall Layers", "ns_walls", ns_wall_layers)
        add_section("Beam Layers", "beams", beam_layers)
        
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=20)
        
        def on_proceed():
            def get_lines(key):
                text = self.layer_inputs[key].get("0.0", "end").strip()
                return [line.strip() for line in text.split('\n') if line.strip()]
                
            self.sel_slabs = get_lines("slabs")
            self.sel_str_walls = get_lines("str_walls")
            self.sel_ns_walls = get_lines("ns_walls")
            self.sel_beams = get_lines("beams")
            
            dialog.destroy()
            self._start_takeoff_thread()
            
        def on_cancel():
            dialog.destroy()
            self.log("Take-off cancelled by user.")
            self._enable_buttons()
            
        proceed_btn = ctk.CTkButton(btn_frame, text="Proceed", command=on_proceed, fg_color="#2FA572", hover_color="#1F7A52")
        proceed_btn.pack(side="right", padx=10)
        
        cancel_btn = ctk.CTkButton(btn_frame, text="Cancel", command=on_cancel, fg_color="gray")
        cancel_btn.pack(side="right")

    def _start_takeoff_thread(self):
        self.log(f"Starting extraction for: {os.path.basename(self.selected_file)}")
        # Run in a separate thread so UI doesn't freeze
        thread = threading.Thread(target=self._process_file)
        thread.start()

    def _process_file(self):
        try:
            dxf_path = self.selected_file
            
            # 1. Extract Slabs
            self.log("Extracting slab geometries and text...")
            report = extract_slabs(dxf_path, allowed_layers=self.sel_slabs)
            
            # 2. Extract Walls
            self.log("Extracting wall geometries...")
            wall_report = extract_walls(dxf_path, str_layers=self.sel_str_walls, ns_layers=self.sel_ns_walls)
            
            # 3. Extract Beams
            self.log("Extracting beam geometries...")
            from beam_extractor import extract_beams
            from beam_visualizer import generate_beam_map
            beam_report = extract_beams(dxf_path, allowed_layers=self.sel_beams, str_wall_layers=self.sel_str_walls)
            
            # 4. Maps
            self.log("Generating visual maps...")
            dxf_dir = os.path.dirname(os.path.abspath(dxf_path))
            base_name = os.path.splitext(os.path.basename(dxf_path))[0]
            
            img_path = os.path.join(dxf_dir, f"{base_name}_slab_map.png")
            render_slab_map(report, img_path)
            
            str_wall_img = os.path.join(dxf_dir, f"{base_name}_str_wall_map.png")
            render_wall_map(wall_report, str_wall_img, is_structural=True)
            
            ns_wall_img = os.path.join(dxf_dir, f"{base_name}_ns_wall_map.png")
            render_wall_map(wall_report, ns_wall_img, is_structural=False)
            
            beam_img = os.path.join(dxf_dir, f"{base_name}_beam_map.png")
            generate_beam_map(beam_report, beam_img)

            # 5. Excel
            self.log("Writing Excel report...")
            excel_path = os.path.join(dxf_dir, f"{base_name}_takeoff_v3.xlsx")
            write_excel_report(report, excel_path, image_path=img_path, 
                               wall_report=wall_report, str_wall_img=str_wall_img, ns_wall_img=ns_wall_img,
                               beam_report=beam_report, beam_img=beam_img)

            # 5. Apply User Height to Excel
            try:
                user_height = float(self.height_var.get())
                from openpyxl import load_workbook
                wb = load_workbook(excel_path)
                if "Structural Walls" in wb.sheetnames:
                    wb["Structural Walls"]["B3"] = user_height
                if "Non-Structural Walls" in wb.sheetnames:
                    wb["Non-Structural Walls"]["B3"] = user_height
                wb.save(excel_path)
            except Exception as he:
                self.log(f"Warning: Could not set custom height in excel: {he}")

            # 7. Summary Output
            self.log("\n--- SUMMARY ---")
            self.log(f"Total Slabs: {report.total_slabs}")
            self.log(f"Total Str Walls: {len(wall_report.structural_walls)}")
            self.log(f"Total NS Walls: {len(wall_report.non_structural_walls)}")
            self.log(f"Total Beams: {len(beam_report.beams)}")
            
            self.log(f"\n✅ SUCCESS!")
            self.log(f"Excel saved to: {excel_path}")

            # 7. Display image in UI (show NS wall as it's the last one)
            self.after(0, self._display_image, str_wall_img)

        except Exception as e:
            self.log(f"\n❌ ERROR: {str(e)}")
        
        finally:
            # Re-enable buttons
            self.after(0, self._enable_buttons)

    def _display_image(self, img_path):
        from PIL import Image
        try:
            pil_image = Image.open(img_path)
            # Resize image to fit the preview panel while maintaining aspect ratio
            target_width, target_height = 450, 500
            pil_image.thumbnail((target_width, target_height))
            
            ctk_img = ctk.CTkImage(light_image=pil_image, dark_image=pil_image, 
                                   size=(pil_image.width, pil_image.height))
            
            self.image_label.configure(image=ctk_img, text="")
            self.image_label.image = ctk_img  # keep a reference
        except Exception as e:
            self.log(f"Failed to load image preview: {e}")

    def _enable_buttons(self):
        self.select_btn.configure(state="normal")
        self.run_btn.configure(state="normal")

if __name__ == "__main__":
    app = TakeoffApp()
    app.mainloop()
