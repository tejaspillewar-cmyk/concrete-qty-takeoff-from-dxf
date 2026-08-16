import sys
import os
import tempfile
import logging

from ezdxf.addons.xqt import QtWidgets as qw, QtCore as qc, QtGui as qg

# Add the 01_slab_takeoff directory to path so we can import its modules
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SLAB_DIR = os.path.join(BASE_DIR, "01_slab_takeoff")
sys.path.insert(0, SLAB_DIR)

from slab_extractor import extract_slabs
from excel_report import write_excel_report
from slab_visualizer import render_slab_map
from wall_extractor import extract_walls
from wall_visualizer import render_wall_map
from beam_extractor import extract_beams
from beam_visualizer import generate_beam_map

import ezdxf
from ezdxf.addons.drawing.qtviewer import CADViewer
from ezdxf.addons.drawing.config import Configuration


# ──────────────────────────────────────────────────────────────────────────────
# Zoomable & Pannable Graphics View for Static Map PNGs
# ──────────────────────────────────────────────────────────────────────────────
class ZoomableImageWidget(qw.QGraphicsView):
    def __init__(self, placeholder_text="Run Take-Off to generate Map", parent=None):
        super().__init__(parent)
        self.placeholder_text = placeholder_text
        self.scene = qw.QGraphicsScene(self)
        self.setScene(self.scene)
        self.pixmap_item = None
        
        # Enable dragging to pan
        self.setDragMode(qw.QGraphicsView.ScrollHandDrag)
        self.setTransformationAnchor(qw.QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(qw.QGraphicsView.AnchorUnderMouse)
        self.setHorizontalScrollBarPolicy(qc.Qt.ScrollBarAsNeeded)
        self.setVerticalScrollBarPolicy(qc.Qt.ScrollBarAsNeeded)
        self.setStyleSheet("background-color: #1a1a2e; border: none;")
        
        self.show_placeholder(self.placeholder_text)

    def show_placeholder(self, text_str):
        self.scene.clear()
        self.pixmap_item = None
        
        # Draw placeholder text centered in scene
        text = self.scene.addText(text_str)
        font = qg.QFont("Segoe UI", 14)
        text.setFont(font)
        text.setDefaultTextColor(qg.QColor("#90a4ae"))
        
        # Center the text
        rect = text.boundingRect()
        text.setPos(-rect.width() / 2, -rect.height() / 2)
        self.setSceneRect(-200, -200, 400, 400)

    def set_image(self, img_path):
        self.scene.clear()
        self.pixmap_item = None
        
        if os.path.exists(img_path):
            pixmap = qg.QPixmap(img_path)
            self.pixmap_item = self.scene.addPixmap(pixmap)
            self.setSceneRect(qc.QRectF(pixmap.rect()))
            
            # Reset scaling/transform matrix to default
            self.resetTransform()
            
            # Fit the visualizer map perfectly into view on load
            self.fitInView(self.pixmap_item, qc.Qt.KeepAspectRatio)
        else:
            self.show_placeholder(f"Error: Map image not found at\n{os.path.basename(img_path)}")

    def wheelEvent(self, event):
        # Zoom with mouse wheel scroll
        zoom_in_factor = 1.15
        zoom_out_factor = 0.85
        
        if event.angleDelta().y() > 0:
            factor = zoom_in_factor
        else:
            factor = zoom_out_factor
            
        self.scale(factor, factor)

    def clear(self):
        self.show_placeholder(self.placeholder_text)


# ──────────────────────────────────────────────────────────────────────────────
# Worker Thread for Async Take-Off Extraction
# ──────────────────────────────────────────────────────────────────────────────
class TakeoffWorker(qc.QThread):
    log_signal = qc.Signal(str)
    finished_signal = qc.Signal(bool, str)

    def __init__(self, dxf_path, floor_height, slab_layers, str_wall_layers, ns_wall_layers, beam_layers):
        super().__init__()
        self.dxf_path = dxf_path
        self.floor_height = floor_height
        self.slab_layers = slab_layers
        self.str_wall_layers = str_wall_layers
        self.ns_wall_layers = ns_wall_layers
        self.beam_layers = beam_layers

    def run(self):
        try:
            dxf_path = self.dxf_path
            
            # 1. Slabs
            self.log_signal.emit("Extracting slab geometries and text...")
            report = extract_slabs(dxf_path, allowed_layers=self.slab_layers)
            
            # 2. Walls
            self.log_signal.emit("Extracting wall geometries...")
            wall_report = extract_walls(dxf_path, str_layers=self.str_wall_layers, ns_layers=self.ns_wall_layers)
            
            # 2b. Beams
            self.log_signal.emit("Extracting beam geometries...")
            beam_report = extract_beams(dxf_path, allowed_layers=self.beam_layers, str_wall_layers=self.str_wall_layers)
            
            # 3. Maps
            self.log_signal.emit("Generating visual maps...")
            dxf_dir = os.path.dirname(os.path.abspath(dxf_path))
            base_name = os.path.splitext(os.path.basename(dxf_path))[0]
            
            img_path = os.path.join(dxf_dir, f"{base_name}_slab_map.png")
            render_slab_map(report, img_path)
            
            str_wall_img = os.path.join(dxf_dir, f"{base_name}_str_wall_map.png")
            render_wall_map(wall_report, str_wall_img, is_structural=True)
            
            ns_wall_img = os.path.join(dxf_dir, f"{base_name}_ns_wall_map.png")
            render_wall_map(wall_report, ns_wall_img, is_structural=False)
            
            beam_img = os.path.join(dxf_dir, f"{base_name}_beam_map.png")
            generate_beam_map(beam_report, beam_img)

            # 4. Excel Report
            self.log_signal.emit("Writing Excel report...")
            excel_path = os.path.join(dxf_dir, f"{base_name}_takeoff_v3.xlsx")
            write_excel_report(report, excel_path, image_path=img_path, 
                               wall_report=wall_report, str_wall_img=str_wall_img, ns_wall_img=ns_wall_img,
                               beam_report=beam_report, beam_img=beam_img)

            # 5. Apply User Height to Excel
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_path)
                if "Structural Walls" in wb.sheetnames:
                    wb["Structural Walls"]["B3"] = self.floor_height
                if "Non-Structural Walls" in wb.sheetnames:
                    wb["Non-Structural Walls"]["B3"] = self.floor_height
                wb.save(excel_path)
            except Exception as he:
                self.log_signal.emit(f"Warning: Could not set custom height in Excel: {he}")

            # 6. Success message
            summary = (
                f"\n--- TAKE-OFF COMPLETED ---\n"
                f"Slabs found: {report.total_slabs}\n"
                f"Structural Walls: {len(wall_report.structural_walls)}\n"
                f"Non-Structural Walls: {len(wall_report.non_structural_walls)}\n"
                f"Beams: {len(beam_report.beams)}\n\n"
                f"Excel Output saved to:\n{excel_path}"
            )
            self.finished_signal.emit(True, summary)
            
        except Exception as e:
            self.finished_signal.emit(False, str(e))


# ──────────────────────────────────────────────────────────────────────────────
# Layer Confirmation Dialog
# ──────────────────────────────────────────────────────────────────────────────
class LayerConfirmDialog(qw.QDialog):
    def __init__(self, slab_layers, str_wall_layers, ns_wall_layers, beam_layers, all_layers, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Confirm Detected Layers")
        self.resize(600, 700)
        
        layout = qw.QVBoxLayout(self)
        
        # Scroll area for layers
        scroll = qw.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = qw.QWidget()
        scroll_layout = qw.QVBoxLayout(scroll_content)
        
        self.lists = {}
        self.all_layers = all_layers
        
        def add_section(title, key, initial_layers):
            lbl = qw.QLabel(title)
            font = qg.QFont()
            font.setBold(True)
            font.setPointSize(11)
            lbl.setFont(font)
            
            header_layout = qw.QHBoxLayout()
            header_layout.addWidget(lbl)
            
            add_btn = qw.QPushButton("+")
            add_btn.setFixedSize(30, 30)
            
            rm_btn = qw.QPushButton("-")
            rm_btn.setFixedSize(30, 30)
            
            header_layout.addStretch()
            header_layout.addWidget(add_btn)
            header_layout.addWidget(rm_btn)
            
            scroll_layout.addLayout(header_layout)
            
            list_widget = qw.QListWidget()
            list_widget.addItems(initial_layers)
            list_widget.setFixedHeight(120)
            scroll_layout.addWidget(list_widget)
            
            self.lists[key] = list_widget
            
            def on_add():
                layer, ok = qw.QInputDialog.getItem(self, "Add Layer", "Select layer to add:", self.all_layers, 0, False)
                if ok and layer:
                    items = [list_widget.item(i).text() for i in range(list_widget.count())]
                    if layer not in items:
                        list_widget.addItem(layer)
            
            def on_rm():
                for item in list_widget.selectedItems():
                    list_widget.takeItem(list_widget.row(item))
                    
            add_btn.clicked.connect(on_add)
            rm_btn.clicked.connect(on_rm)
            
        add_section("Slab Layers", "slabs", slab_layers)
        add_section("Structural Wall Layers", "str_walls", str_wall_layers)
        add_section("Non-Structural Wall Layers", "ns_walls", ns_wall_layers)
        add_section("Beam Layers", "beams", beam_layers)
        
        scroll_layout.addStretch()
        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)
        
        # Buttons
        btn_box = qw.QDialogButtonBox()
        btn_box.addButton("Proceed", qw.QDialogButtonBox.AcceptRole)
        btn_box.addButton("Cancel", qw.QDialogButtonBox.RejectRole)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def get_selections(self):
        def get_items(key):
            lw = self.lists[key]
            return [lw.item(i).text() for i in range(lw.count())]
        return (
            get_items("slabs"),
            get_items("str_walls"),
            get_items("ns_walls"),
            get_items("beams")
        )

# ──────────────────────────────────────────────────────────────────────────────
# Main Application Window
# ──────────────────────────────────────────────────────────────────────────────
class TakeoffApp(qw.QMainWindow):
    def __init__(self):
        super().__init__()
        self.selected_file = None
        self.worker = None

        self.setWindowTitle("Slab & Wall Quantity Take-Off")
        self.resize(1400, 850)
        self.setup_ui()
        self.apply_styles()

    def setup_ui(self):
        # Central splitter (Left panel: Controls, Right panel: Tabs with Viewer/Maps)
        self.main_splitter = qw.QSplitter(qc.Qt.Horizontal, self)
        self.setCentralWidget(self.main_splitter)

        # ── LEFT PANEL: Controls ──────────────────────────────────────────────
        self.left_widget = qw.QWidget()
        self.left_layout = qw.QVBoxLayout(self.left_widget)
        self.left_layout.setContentsMargins(20, 20, 20, 20)
        self.left_layout.setSpacing(15)

        # Header Title
        self.title_label = qw.QLabel("Slab & Wall Take-Off")
        self.title_label.setObjectName("TitleLabel")
        self.left_layout.addWidget(self.title_label)

        self.subtitle_label = qw.QLabel("Extract structural quantities and visualize plans.")
        self.subtitle_label.setObjectName("SubtitleLabel")
        self.left_layout.addWidget(self.subtitle_label)

        # File Selection Frame
        self.file_group = qw.QGroupBox("CAD Document")
        self.file_layout = qw.QHBoxLayout(self.file_group)
        self.file_layout.setContentsMargins(10, 15, 10, 10)
        
        self.select_btn = qw.QPushButton("Browse DXF")
        self.select_btn.clicked.connect(self.select_file)
        self.file_layout.addWidget(self.select_btn)

        self.file_path_edit = qw.QLineEdit("No file selected")
        self.file_path_edit.setReadOnly(True)
        self.file_path_edit.setObjectName("FilePathEdit")
        self.file_layout.addWidget(self.file_path_edit)

        self.left_layout.addWidget(self.file_group)

        # Settings Frame
        self.settings_group = qw.QGroupBox("Take-Off Parameters")
        self.settings_layout = qw.QFormLayout(self.settings_group)
        self.settings_layout.setContentsMargins(10, 15, 10, 10)
        self.settings_layout.setVerticalSpacing(10)

        self.height_label = qw.QLabel("Floor-to-Floor Height (m):")
        self.height_edit = qw.QLineEdit("3.0")
        self.height_edit.setFixedWidth(80)
        self.settings_layout.addRow(self.height_label, self.height_edit)

        self.left_layout.addWidget(self.settings_group)

        # Action Buttons
        self.run_btn = qw.QPushButton("Run Take-Off")
        self.run_btn.setObjectName("RunButton")
        self.run_btn.clicked.connect(self.run_takeoff)
        self.left_layout.addWidget(self.run_btn)

        # Console Logs
        self.log_label = qw.QLabel("Process Log Console:")
        self.log_label.setObjectName("SectionHeader")
        self.left_layout.addWidget(self.log_label)

        self.log_textbox = qw.QTextEdit()
        self.log_textbox.setReadOnly(True)
        self.log_textbox.setObjectName("LogConsole")
        self.log_textbox.append("Welcome! Browse and select a DXF file to inspect geometry.")
        self.left_layout.addWidget(self.log_textbox)

        self.main_splitter.addWidget(self.left_widget)

        # ── RIGHT PANEL: Tabs with CAD Viewer & Zoomable Map Previews ─────────
        self.right_tabs = qw.QTabWidget()
        
        # Tab 1: Interactive CAD
        config = Configuration(lineweight_scaling=0)
        self.viewer = CADViewer.from_config(config)
        self.viewer.menuBar().hide() # Hide standard CADViewer menu bar for clean look
        self.right_tabs.addTab(self.viewer, "Interactive CAD View")
        
        # Tab 2: Slab Map (Zoomable)
        self.slab_view = ZoomableImageWidget("Run Take-Off to generate Slab Map")
        self.right_tabs.addTab(self.slab_view, "Slab Map")
        
        # Tab 3: Structural Wall Map (Zoomable)
        self.str_wall_view = ZoomableImageWidget("Run Take-Off to generate Structural Wall Map")
        self.right_tabs.addTab(self.str_wall_view, "Structural Wall Map")
        
        # Tab 4: Non-Structural Wall Map (Zoomable)
        self.ns_wall_view = ZoomableImageWidget("Run Take-Off to generate Non-Structural Wall Map")
        self.right_tabs.addTab(self.ns_wall_view, "Non-Structural Wall Map")

        # Tab 5: Beam Map (Zoomable)
        self.beam_view = ZoomableImageWidget("Run Take-Off to generate Beam Map")
        self.right_tabs.addTab(self.beam_view, "Beam Map")

        self.main_splitter.addWidget(self.right_tabs)

        # Set ratio for splitter: 30% controls, 70% Right tab content
        self.main_splitter.setSizes([350, 1050])

    def apply_styles(self):
        # Modern Dark Blue Slate Theme matching professional CAD design
        self.setStyleSheet("""
            QMainWindow {
                background-color: #1e272c;
            }
            QWidget {
                color: #eceff1;
                font-family: "Segoe UI", Arial, sans-serif;
                font-size: 13px;
            }
            QSplitter::handle {
                background-color: #2c3b42;
                width: 2px;
            }
            #TitleLabel {
                font-size: 22px;
                font-weight: bold;
                color: #2FA572;
                margin-top: 10px;
            }
            #SubtitleLabel {
                font-size: 13px;
                color: #90a4ae;
                margin-bottom: 10px;
            }
            #SectionHeader {
                font-weight: bold;
                color: #90a4ae;
                margin-top: 5px;
            }
            QGroupBox {
                border: 1px solid #37474f;
                border-radius: 6px;
                margin-top: 15px;
                font-weight: bold;
                color: #90a4ae;
                background-color: #263238;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QPushButton {
                background-color: #37474f;
                border: 1px solid #4f5b66;
                border-radius: 4px;
                padding: 6px 12px;
                font-weight: bold;
                color: #ffffff;
            }
            QPushButton:hover {
                background-color: #455a64;
                border-color: #78909c;
            }
            QPushButton:pressed {
                background-color: #263238;
            }
            #RunButton {
                background-color: #2FA572;
                border: 1px solid #1f7a52;
                color: white;
                font-size: 15px;
                padding: 10px;
                border-radius: 6px;
            }
            #RunButton:hover {
                background-color: #38b883;
                border-color: #2FA572;
            }
            #RunButton:disabled {
                background-color: #37474f;
                color: #78909c;
                border-color: #37474f;
            }
            QLineEdit {
                background-color: #1c2529;
                border: 1px solid #37474f;
                border-radius: 4px;
                padding: 5px;
                color: #ffffff;
            }
            QLineEdit:focus {
                border: 1px solid #2FA572;
            }
            #FilePathEdit {
                background-color: #1c2529;
                color: #cfd8dc;
            }
            #LogConsole {
                background-color: #10171a;
                border: 1px solid #263238;
                border-radius: 6px;
                color: #81c784;
                font-family: Consolas, "Courier New", monospace;
                font-size: 12px;
            }
            QTabWidget::pane {
                border: 1px solid #37474f;
                background-color: #1c2529;
                border-radius: 4px;
            }
            QTabBar::tab {
                background-color: #263238;
                color: #90a4ae;
                padding: 8px 16px;
                border: 1px solid #37474f;
                border-bottom-color: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                font-weight: bold;
            }
            QTabBar::tab:hover {
                background-color: #37474f;
            }
            QTabBar::tab:selected {
                background-color: #1c2529;
                color: #ffffff;
                border-bottom-color: #1c2529;
            }
        """)

    def log(self, message):
        self.log_textbox.append(message)
        self.log_textbox.ensureCursorVisible()

    def select_file(self):
        filename, _ = qw.QFileDialog.getOpenFileName(
            self,
            "Select DXF Drawing",
            "",
            "DXF Files (*.dxf);;All Files (*)"
        )
        if filename:
            self.selected_file = filename
            self.file_path_edit.setText(os.path.basename(filename))
            self.log(f"\n📂 Selected CAD File:\n{filename}")
            
            # Reset map displays
            self.slab_view.clear()
            self.str_wall_view.clear()
            self.ns_wall_view.clear()
            self.right_tabs.setCurrentIndex(0)
            
            # Load into embedded CADViewer
            self.log("Loading DXF structure into interactive viewer...")
            try:
                self.viewer.load_file(filename)
                self.log("Interactive CAD Preview is ready. You can pan, zoom, and toggle layers.")
            except Exception as e:
                self.log(f"⚠️ Failed to render CAD view: {e}")

    def run_takeoff(self):
        if not self.selected_file:
            qw.QMessageBox.warning(self, "No File", "Please select a DXF file first.")
            return

        try:
            floor_height = float(self.height_edit.text())
        except ValueError:
            qw.QMessageBox.warning(self, "Invalid Value", "Please enter a numeric height.")
            return

        # Disable buttons during execution
        self.select_btn.setEnabled(False)
        self.run_btn.setEnabled(False)

        self.log("\n" + "="*50)
        self.log(f"Scanning layers for: {os.path.basename(self.selected_file)}")
        
        try:
            import ezdxf
            doc = ezdxf.readfile(self.selected_file)
            all_layers = [layer.dxf.name for layer in doc.layers]
        except Exception as e:
            self.log(f"Failed to read DXF layers: {e}")
            self.select_btn.setEnabled(True)
            self.run_btn.setEnabled(True)
            return
            
        slab_layers = [l for l in all_layers if l.upper().startswith("STR-SLAB-REG")]
        str_wall_layers = [l for l in all_layers if "STR-WALL-REG" in l.upper()]
        ns_wall_layers = [l for l in all_layers if "NS-WALL-REG" in l.upper()]
        beam_layers = [l for l in all_layers if "BEAM" in l.upper()]
        
        dialog = LayerConfirmDialog(slab_layers, str_wall_layers, ns_wall_layers, beam_layers, all_layers, self)
        if dialog.exec() == qw.QDialog.Accepted:
            sel_slabs, sel_str_walls, sel_ns_walls, sel_beams = dialog.get_selections()
            self.log(f"Starting Take-Off Analysis...")
            # Start Worker Thread
            self.worker = TakeoffWorker(self.selected_file, floor_height, sel_slabs, sel_str_walls, sel_ns_walls, sel_beams)
            self.worker.log_signal.connect(self.log)
            self.worker.finished_signal.connect(self.on_takeoff_finished)
            self.worker.start()
        else:
            self.log("Take-off cancelled by user.")
            self.select_btn.setEnabled(True)
            self.run_btn.setEnabled(True)

    def on_takeoff_finished(self, success, message):
        if success:
            self.log(message)
            self.log("\n🚀 Quantity takeoff process completed successfully!")
            
            # Load the generated visualizer maps into respective tabs
            dxf_dir = os.path.dirname(os.path.abspath(self.selected_file))
            base_name = os.path.splitext(os.path.basename(self.selected_file))[0]
            
            slab_img = os.path.join(dxf_dir, f"{base_name}_slab_map.png")
            str_wall_img = os.path.join(dxf_dir, f"{base_name}_str_wall_map.png")
            ns_wall_img = os.path.join(dxf_dir, f"{base_name}_ns_wall_map.png")
            beam_img = os.path.join(dxf_dir, f"{base_name}_beam_map.png")
            
            self.slab_view.set_image(slab_img)
            self.str_wall_view.set_image(str_wall_img)
            self.ns_wall_view.set_image(ns_wall_img)
            self.beam_view.set_image(beam_img)
            
            # Auto-switch to the Slab Map tab so the user sees the output immediately
            self.right_tabs.setCurrentIndex(1)
            
            qw.QMessageBox.information(self, "Take-Off Complete", "Slab and Wall Quantities extracted successfully.")
        else:
            self.log(f"\n❌ FAILED:\n{message}")
            qw.QMessageBox.critical(self, "Take-Off Error", f"Extraction failed:\n{message}")

        # Re-enable controls
        self.select_btn.setEnabled(True)
        self.run_btn.setEnabled(True)


if __name__ == "__main__":
    app = qw.QApplication(sys.argv)
    window = TakeoffApp()
    window.show()
    sys.exit(app.exec())
